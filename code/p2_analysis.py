"""
P2 analyses for the HCV genotype 6 / NS5A-28 manuscript (F1000Research).

(A) Firth penalised logistic regression, outcome ~ subtype group + regimen
(B) Exact Clopper-Pearson 95% CIs for every subtype and regimen failure rate
(C) Prevalence of a position-28 polymorphism by subtype

Plus the stratified Fisher tests that carry Results R2-R6 and the numbers
behind the three panels of the new central figure.

Input : master_162.csv
Output: P2_results.xlsx  (one sheet per table)  +  console log
"""
import numpy as np
import pandas as pd
from scipy import stats
from statsmodels.stats.proportion import proportion_confint

from firth import firth_logit

SRC = '/mnt/user-data/uploads/master_162.csv'
OUT = '/mnt/user-data/outputs/P2_results.xlsx'

df = pd.read_csv(SRC)
df['fail'] = (df['outcome'] == 'Non SVR').astype(int)

# ---------------------------------------------------------------- helpers
def cp(k, n, dec=1):
    """Failure rate with exact Clopper-Pearson 95% CI, as strings."""
    if n == 0:
        return '-', '-'
    lo, hi = proportion_confint(k, n, alpha=0.05, method='beta')
    return f'{100*k/n:.{dec}f}', f'{100*lo:.{dec}f}-{100*hi:.{dec}f}'


def fisher(a, b, c, d):
    """a/b = events/non-events in group 1; c/d in group 2."""
    orr, p = stats.fisher_exact([[a, b], [c, d]])
    return orr, p


def rate_row(sub, label):
    n, k = len(sub), int(sub['fail'].sum())
    pct, ci = cp(k, n)
    return {'Group': label, 'n': n, 'Non-SVR': k, 'Failure rate (%)': pct,
            '95% CI (Clopper-Pearson)': ci}


# ---------------------------------------------------- derived variables
def subtype_group(r):
    if r['subtype'] == '6r':
        return '6r'
    if r['GT'] == 6:
        return 'GT6 other'
    if r['subtype'] == '1l':
        return '1l'
    if r['GT'] == 1:
        return 'GT1 other'
    if r['GT'] == 2:
        return 'GT2'
    return 'GT4'


df['subtype_group'] = df.apply(subtype_group, axis=1)
df['regimen_group'] = df['regimen'].where(
    df['regimen'].isin(['SOF-VEL', 'SOF-DCV', 'SOF-LDV']), 'Other regimens')

# --- reference correction for subtype 6r --------------------------------
# geno2pheno[HCV] scores each sequence against its NEAREST subtype reference,
# so the "wild-type" letter in a call depends on which reference was picked.
# The ICTV reference for 6r is EU408328 (isolate QC245), which carries G at
# NS5A position 28 - hence the G28T / G28A / G28K calls in 19 of 21 sequences.
# Two 6r sequences were instead scored against DQ835760, the 6f reference,
# which carries A at position 28; their "A28T" calls report the SAME observed
# residue (T) against the wrong reference. Re-score them against the 6r
# reference so that every 6r sequence is read against EU408328.
df['ref_6f_mismatch'] = (df['subtype'] == '6r') & (df['ns5a_28'] == 'A28T')
df['ns5a_28_raw'] = df['ns5a_28']
df.loc[df['ref_6f_mismatch'], 'ns5a_28'] = 'G28T'

# position-28 polymorphism (relative to the assigned subtype reference)
df['sub28'] = df['ns5a_28'].notna()
# residue actually present at 28 (second half of e.g. G28T)
df['res28'] = df['ns5a_28'].str[-1].where(df['sub28'])

writer_sheets = {}
log = []


def emit(name, frame):
    writer_sheets[name] = frame
    log.append(f'\n===== {name} =====')
    log.append(frame.to_string(index=False))


# ============================================================ TABLE 1
# Cohort: subtype x outcome with exact CIs, then regimen x outcome
rows = []
for gt in sorted(df['GT'].unique()):
    sub = df[df['GT'] == gt]
    rows.append({**rate_row(sub, f'Genotype {gt}'), 'Level': 'Genotype'})
    for st in sorted(sub['subtype'].unique()):
        s2 = sub[sub['subtype'] == st]
        rows.append({**rate_row(s2, f'  {st}'), 'Level': 'Subtype'})
rows.append({**rate_row(df, 'All'), 'Level': 'Total'})
t1a = pd.DataFrame(rows)[['Level', 'Group', 'n', 'Non-SVR',
                          'Failure rate (%)', '95% CI (Clopper-Pearson)']]
emit('T1a_subtype', t1a)

rows = [rate_row(df[df['regimen'] == r], r) for r in sorted(df['regimen'].unique())]
rows.append(rate_row(df, 'All'))
t1b = pd.DataFrame(rows)
chi2, pchi, _, _ = stats.chi2_contingency(
    pd.crosstab(df['regimen'], df['outcome']).values)
t1b.loc[len(t1b)] = {'Group': f'Chi-square regimen x outcome: P = {pchi:.3g}',
                     'n': '', 'Non-SVR': '', 'Failure rate (%)': '',
                     '95% CI (Clopper-Pearson)': ''}
emit('T1b_regimen', t1b)

# subtype x regimen cross-tabulation (declares the confounding structure)
t1c = pd.crosstab(df['subtype'], df['regimen']).reset_index()
emit('T1c_subtype_by_regimen', t1c)

# ============================================================ R2 headline
g6dcv = df[(df['GT'] == 6) & (df['regimen'] == 'SOF-DCV')]
r6 = g6dcv[g6dcv['subtype'] == '6r']
ro = g6dcv[g6dcv['subtype'] != '6r']
a, b = int(r6['fail'].sum()), len(r6) - int(r6['fail'].sum())
c, d = int(ro['fail'].sum()), len(ro) - int(ro['fail'].sum())
orr, p_r2 = fisher(a, b, c, d)
rows = [rate_row(r6, 'Subtype 6r'), rate_row(ro, 'Other genotype 6 subtypes')]
for st in sorted(g6dcv['subtype'].unique()):
    rows.append(rate_row(g6dcv[g6dcv['subtype'] == st], f'  {st}'))
t2 = pd.DataFrame(rows)
t2.loc[len(t2)] = {'Group': f'6r vs other GT6 (SOF/DCV only): Fisher P = {p_r2:.3g}, '
                            f'unadjusted OR = {orr:.1f}',
                   'n': '', 'Non-SVR': '', 'Failure rate (%)': '',
                   '95% CI (Clopper-Pearson)': ''}
emit('T2_R2_GT6_on_SOFDCV', t2)

# ============================================================ R3/R4 NS5A-28
g6 = df[df['GT'] == 6]
rows = []
for st in sorted(g6['subtype'].unique()):
    s = g6[g6['subtype'] == st]
    n = len(s)
    n28 = int(s['sub28'].sum())
    pct28, ci28 = cp(n28, n)
    res = ', '.join(f'{k} {v}' for k, v in
                    s.loc[s['sub28'], 'ns5a_28'].value_counts().items()) or '-'
    n62 = int(s['ns5a_62'].notna().sum())
    n30 = int(s['ns5a_30'].notna().sum())
    n31 = int(s['ns5a_31'].notna().sum())
    n93 = int(s['ns5a_93'].notna().sum())
    rows.append({'Subtype': st, 'n': n, 'Non-SVR': int(s['fail'].sum()),
                 'Position 28 substitution, n': n28,
                 '% (95% CI)': f'{pct28} ({ci28})',
                 'Residues at 28': res,
                 'Position 30, n': n30, 'Position 31, n': n31,
                 'Position 62, n': n62, 'Position 93, n': n93})
t3 = pd.DataFrame(rows)
emit('T3_NS5A_map_GT6', t3)

# any-28 vs no-28 within SOF/DCV-treated GT6
y28 = g6dcv[g6dcv['sub28']]
n28g = g6dcv[~g6dcv['sub28']]
a, b = int(y28['fail'].sum()), len(y28) - int(y28['fail'].sum())
c, d = int(n28g['fail'].sum()), len(n28g) - int(n28g['fail'].sum())
or28, p28 = fisher(a, b, c, d)

# residue identity, among 28-substituted GT6 on SOF/DCV
y28 = y28.copy()
y28['res_class'] = np.where(y28['res28'].isin(['T', 'K']), 'T or K',
                            np.where(y28['res28'] == 'A', 'A', 'other'))
tk = y28[y28['res_class'] == 'T or K']
al = y28[y28['res_class'] == 'A']
a, b = int(tk['fail'].sum()), len(tk) - int(tk['fail'].sum())
c, d = int(al['fail'].sum()), len(al) - int(al['fail'].sum())
or_res, p_res = fisher(a, b, c, d)

rows = [rate_row(y28, 'Any position-28 substitution'),
        rate_row(n28g, 'No position-28 substitution'),
        rate_row(tk, '  Residue T or K'),
        rate_row(al, '  Residue A'),
        rate_row(y28[y28['res_class'] == 'other'], '  Other residue')]
t4 = pd.DataFrame(rows)
t4.loc[len(t4)] = {'Group': f'Any 28 vs none (GT6 on SOF/DCV): Fisher P = {p28:.4g}, OR = {or28:.2f}',
                   'n': '', 'Non-SVR': '', 'Failure rate (%)': '', '95% CI (Clopper-Pearson)': ''}
t4.loc[len(t4)] = {'Group': f'T/K vs A: Fisher P = {p_res:.4g}, OR = {or_res:.2f}',
                   'n': '', 'Non-SVR': '', 'Failure rate (%)': '', '95% CI (Clopper-Pearson)': ''}
emit('T4_R3R4_position28', t4)

# ---- confounding checks a reviewer will ask for -------------------------
# (i) every 6r carries a 28 substitution, so "any 28" and "6r" cannot be
#     separated in this cohort. Test 28 in NON-6r GT6 on SOF/DCV.
non6r = g6dcv[g6dcv['subtype'] != '6r']
n28y = non6r[non6r['sub28']]
n28n = non6r[~non6r['sub28']]
a, b = int(n28y['fail'].sum()), len(n28y) - int(n28y['fail'].sum())
c, d = int(n28n['fail'].sum()), len(n28n) - int(n28n['fail'].sum())
or_n, p_n = fisher(a, b, c, d)

# (ii) residue identity restricted to 6r, excluding the two sequences whose
#      geno2pheno call maps to the divergent DQ835760 reference (the A28T calls)
r6_clean = g6dcv[(g6dcv['subtype'] == '6r') & (~g6dcv['ref_6f_mismatch'])].copy()
tk2 = r6_clean[r6_clean['ns5a_28'].isin(['G28T', 'G28K'])]
a2 = r6_clean[r6_clean['ns5a_28'] == 'G28A']
a, b = int(tk2['fail'].sum()), len(tk2) - int(tk2['fail'].sum())
c, d = int(a2['fail'].sum()), len(a2) - int(a2['fail'].sum())
or_c, p_c = fisher(a, b, c, d)

rows = [rate_row(n28y, 'Non-6r GT6 on SOF/DCV, 28 substitution'),
        rate_row(n28n, 'Non-6r GT6 on SOF/DCV, no 28 substitution'),
        rate_row(tk2, 'Subtype 6r only, G28T or G28K'),
        rate_row(a2, 'Subtype 6r only, G28A')]
t4b = pd.DataFrame(rows)
t4b.loc[len(t4b)] = {'Group': f'28 substitution within NON-6r GT6: Fisher P = {p_n:.3g}, OR = {or_n:.2f}',
                     'n': '', 'Non-SVR': '', 'Failure rate (%)': '', '95% CI (Clopper-Pearson)': ''}
t4b.loc[len(t4b)] = {'Group': f'SENSITIVITY - G28T/G28K vs G28A within 6r, dropping the two sequences '
                              f'geno2pheno had scored against the 6f reference DQ835760 instead of the '
                              f'6r reference EU408328: Fisher P = {p_c:.3g}, OR = {or_c:.2f}',
                     'n': '', 'Non-SVR': '', 'Failure rate (%)': '', '95% CI (Clopper-Pearson)': ''}
t4b.loc[len(t4b)] = {'Group': 'REFERENCE CORRECTION: those two sequences carry T at position 28 and were '
                              'reported as "A28T" only because DQ835760 carries A there; against the 6r '
                              'reference EU408328 (G at 28) the call is G28T. They are therefore analysed '
                              'in the G28T group, not as a separate residue.',
                     'n': '', 'Non-SVR': '', 'Failure rate (%)': '', '95% CI (Clopper-Pearson)': ''}
t4b.loc[len(t4b)] = {'Group': 'NOTE: all 21 subtype-6r sequences carry a position-28 substitution, '
                              'so "any 28" and "6r" are collinear in this cohort and cannot be '
                              'separated statistically. All 21 are on SOF/DCV, so the residue-identity '
                              'comparison is within subtype AND within regimen.',
                     'n': '', 'Non-SVR': '', 'Failure rate (%)': '', '95% CI (Clopper-Pearson)': ''}
emit('T4b_confounding_checks', t4b)

# ============================================================ R5 GT2d contrast
d2 = df[df['subtype'] == '2d']
rows = [rate_row(d2, 'Subtype 2d (all)')]
for rg in sorted(d2['regimen'].unique()):
    rows.append(rate_row(d2[d2['regimen'] == rg], f'  {rg}'))
t5 = pd.DataFrame(rows)
t5.loc[len(t5)] = {'Group': f'2d carrying a position-28 substitution: '
                            f'{int(d2["sub28"].sum())}/{len(d2)} '
                            f'({", ".join(d2.loc[d2["sub28"], "ns5a_28"].unique())}); '
                            f'position 31: {int(d2["ns5a_31"].notna().sum())}/{len(d2)}',
                   'n': '', 'Non-SVR': '', 'Failure rate (%)': '', '95% CI (Clopper-Pearson)': ''}
emit('T5_R5_GT2d', t5)

# ============================================================ R6 NS3 in GT6
g6ns3 = g6[g6['covered_ns3']]
rows = []
for pos, col in [('36', 'ns3_36'), ('56', 'ns3_56'), ('80', 'ns3_80'),
                 ('155', 'ns3_155'), ('156', 'ns3_156'), ('168', 'ns3_168')]:
    s = g6ns3[g6ns3[col].notna()] if col in g6ns3 else g6ns3.iloc[0:0]
    subs = ', '.join(f'{k} {v}' for k, v in s[col].value_counts().items()) if len(s) else 'none'
    n = len(s)
    pct, ci = cp(n, len(g6ns3))
    rows.append({'NS3 position': pos, 'n with substitution': n,
                 '% of GT6 with NS3 data (95% CI)': f'{pct} ({ci})',
                 'Substitutions observed': subs,
                 'n non-SVR among them': int(s['fail'].sum()) if n else 0,
                 'Subtypes': ', '.join(sorted(s['subtype'].unique())) if n else '-'})
t6 = pd.DataFrame(rows)
emit('T6_R6_NS3_GT6', t6)

# ============================================================ (C) 28 by subtype, whole cohort
rows = []
for st in sorted(df['subtype'].unique()):
    s = df[df['subtype'] == st]
    n = len(s)
    k = int(s['sub28'].sum())
    pct, ci = cp(k, n)
    rows.append({'Subtype': st, 'Genotype': int(s['GT'].iloc[0]), 'n': n,
                 'With position-28 substitution': k,
                 'Prevalence (%)': pct, '95% CI': ci,
                 'Residues': ', '.join(f'{a} {b}' for a, b in
                                       s.loc[s['sub28'], 'ns5a_28'].value_counts().items()) or '-',
                 'Non-SVR': int(s['fail'].sum())})
t7 = pd.DataFrame(rows)
emit('T7_pos28_prevalence_all', t7)

# ============================================================ (A) Firth models
def design(frame, cols, ref):
    """Build an intercept + dummy design matrix with explicit reference levels."""
    X = [np.ones(len(frame))]
    names = ['Intercept']
    for c in cols:
        levels = [l for l in sorted(frame[c].unique()) if l != ref[c]]
        for l in levels:
            X.append((frame[c] == l).astype(float).values)
            names.append(f'{c}: {l} (vs {ref[c]})')
    return np.column_stack(X), names


def firth_frame(frame, cols, ref, title):
    X, names = design(frame, cols, ref)
    r = firth_logit(X, frame['fail'].values, names)
    out = pd.DataFrame({
        'Term': r['names'],
        'Adjusted OR': [f'{v:.2f}' for v in r['or']],
        '95% CI (profile penalised likelihood)':
            [f'{lo:.2f}-{hi:.2f}' if np.isfinite(lo) and np.isfinite(hi) else 'not estimable'
             for lo, hi in zip(r['ci_low'], r['ci_high'])],
        'P (penalised LR test)': [f'{v:.3g}' for v in r['p']]})
    out.loc[len(out)] = {'Term': f'{title}: n = {r["n"]}, events = {r["events"]}',
                         'Adjusted OR': '', '95% CI (profile penalised likelihood)': '',
                         'P (penalised LR test)': ''}
    return out


m1 = firth_frame(df, ['subtype_group', 'regimen_group'],
                 {'subtype_group': 'GT6 other', 'regimen_group': 'SOF-VEL'},
                 'Model 1, full cohort')
emit('T8_Firth_model1', m1)

g6dcv2 = g6dcv.copy()
g6dcv2['is6r'] = np.where(g6dcv2['subtype'] == '6r', '6r', 'other GT6')
m2 = firth_frame(g6dcv2, ['is6r'], {'is6r': 'other GT6'},
                 'Model 2, genotype 6 on SOF/DCV only')
emit('T9_Firth_model2_6r', m2)

g6dcv2['has28'] = np.where(g6dcv2['sub28'], 'position-28 substitution', 'none')
m3 = firth_frame(g6dcv2, ['has28'], {'has28': 'none'},
                 'Model 3, genotype 6 on SOF/DCV only')
emit('T10_Firth_model3_pos28', m3)

# ============================================================ figure data
fig = []
for st in sorted(g6dcv['subtype'].unique()):
    s = g6dcv[g6dcv['subtype'] == st]
    k, n = int(s['fail'].sum()), len(s)
    lo, hi = proportion_confint(k, n, 0.05, 'beta')
    fig.append({'Panel': 'A', 'Category': st, 'n': n, 'events': k,
                'rate': round(100 * k / n, 1), 'lo': round(100 * lo, 1),
                'hi': round(100 * hi, 1)})
for lab, s in [('Position-28 substitution', g6dcv[g6dcv['sub28']]),
               ('No substitution at 28', g6dcv[~g6dcv['sub28']])]:
    k, n = int(s['fail'].sum()), len(s)
    lo, hi = proportion_confint(k, n, 0.05, 'beta')
    fig.append({'Panel': 'B', 'Category': lab, 'n': n, 'events': k,
                'rate': round(100 * k / n, 1), 'lo': round(100 * lo, 1),
                'hi': round(100 * hi, 1)})
y28b = g6dcv[g6dcv['sub28']].copy()
y28b['res_class'] = np.where(y28b['res28'].isin(['T', 'K']), 'T or K',
                             np.where(y28b['res28'] == 'A', 'A', 'Other residue'))
for lab in ['T or K', 'A', 'Other residue']:
    s = y28b[y28b['res_class'] == lab]
    if len(s) == 0:
        continue
    k, n = int(s['fail'].sum()), len(s)
    lo, hi = proportion_confint(k, n, 0.05, 'beta')
    fig.append({'Panel': 'C', 'Category': lab, 'n': n, 'events': k,
                'rate': round(100 * k / n, 1), 'lo': round(100 * lo, 1),
                'hi': round(100 * hi, 1)})
figd = pd.DataFrame(fig)
emit('T11_figure_data', figd)

# ============================================================ write
with pd.ExcelWriter(OUT, engine='openpyxl') as xw:
    for name, frame in writer_sheets.items():
        frame.to_excel(xw, sheet_name=name[:31], index=False)

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
wb = load_workbook(OUT)
for ws in wb.worksheets:
    for cell in ws[1]:
        cell.font = Font(name='Arial', size=10, bold=True)
        cell.alignment = Alignment(vertical='center', wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name='Arial', size=10)
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 55)
    ws.freeze_panes = 'A2'
wb.save(OUT)

print('\n'.join(log))
print(f'\nWritten: {OUT}')
